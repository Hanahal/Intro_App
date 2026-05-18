import streamlit as st
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import math
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.patches import FancyArrowPatch, Rectangle, Polygon, FancyBboxPatch
from matplotlib.lines import Line2D
from scipy.stats import truncnorm
import io
import datetime
import warnings
warnings.filterwarnings("ignore")

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, PageBreak, Image as RLImage,
)

# ─────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(page_title="CRF Analysis", page_icon="⛏️",
                   layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
section[data-testid="stSidebar"] {display: none !important;}
div[data-testid="stSidebarNav"]  {display: none !important;}

@import url('https://fonts.googleapis.com/css2?family=Lato:wght@300;400;700;900&display=swap');

html, body, [class*="css"], .stApp {
    font-family: 'Lato', sans-serif !important;
    font-size: 15px !important;
    color: #f5e6d0 !important;
}
h1, h2 {
    font-family: 'Lato', sans-serif !important;
    font-weight: 900 !important;
    font-size: 18px !important;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    color: #ffb830 !important;
}
h3, h4 {
    font-family: 'Lato', sans-serif !important;
    font-weight: 700 !important;
    font-size: 16px !important;
    color: #ffb830 !important;
}
p, li, span, div { font-size: 15px !important; }
.stApp { background: #1a1008 !important; color: #f5e6d0 !important; }

/* Tab styling */
.stTabs [data-baseweb="tab-list"] {
    background: #2a1a08 !important;
    border-radius: 8px 8px 0 0 !important;
    padding: 4px 4px 0 4px !important;
    gap: 2px !important;
}
.stTabs [data-baseweb="tab"] {
    font-family: 'Lato', sans-serif !important;
    font-size: 14px !important;
    font-weight: 700 !important;
    color: #c9a87c !important;
    background: #3a2510 !important;
    border-radius: 6px 6px 0 0 !important;
    padding: 10px 18px !important;
    border: 1px solid #5c3d1e !important;
    border-bottom: none !important;
}
.stTabs [aria-selected="true"] {
    color: #1a1008 !important;
    font-weight: 900 !important;
    background: #ffb830 !important;
    border-color: #ffb830 !important;
}
.stTabs [data-baseweb="tab-panel"] {
    background: #221508 !important;
    border: 1px solid #5c3d1e !important;
    border-radius: 0 4px 4px 4px !important;
    padding: 20px !important;
}

/* Inputs */
.stSelectbox label, .stNumberInput label,
.stRadio label, .stSlider label {
    font-size: 15px !important;
    color: #f5e6d0 !important;
    font-weight: 700 !important;
}
.stSelectbox div[data-baseweb="select"] {
    font-size: 15px !important;
    background: #2a1a08 !important;
    border-color: #8b5e2a !important;
    color: #f5e6d0 !important;
}
input[type="number"] {
    background: #2a1a08 !important;
    color: #f5e6d0 !important;
    border-color: #8b5e2a !important;
}

/* Hero */
.hero-banner {
    background: linear-gradient(135deg, #3a2005 0%, #1a1008 50%, #2a1505 100%);
    border: 1px solid #ffb830;
    border-left: 5px solid #ffb830;
    border-radius: 6px;
    padding: 14px 20px;
    margin-bottom: 16px;
}
.hero-banner h1 { color: #ffb830; font-size: 22px !important; }
.hero-banner p  { color: #d4a87c; font-size: 15px !important; }

/* Table */
.styled-table { width: 100%; border-collapse: collapse; font-size: 15px !important; }
.styled-table th {
    background: #3a2005;
    color: #ffb830;
    padding: 10px 14px;
    text-transform: uppercase;
    font-size: 13px !important;
    letter-spacing: 0.04em;
    border-bottom: 2px solid #ffb830;
}
.styled-table td {
    padding: 9px 14px;
    border-bottom: 1px solid #3a2510;
    color: #f0d8b8;
    font-size: 15px !important;
}
.styled-table tr:hover td { background: #2e1a08; }
.styled-table tr:nth-child(even) td { background: #241208; }

/* Metrics */
[data-testid="metric-container"] {
    background: #2a1808 !important;
    border: 1px solid #ffb830 !important;
    border-radius: 10px !important;
    padding: 16px !important;
}
[data-testid="metric-container"] label {
    color: #c9a87c !important;
    font-size: 13px !important;
    font-weight: 700 !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #ffb830 !important;
    font-size: 1.6rem !important;
    font-weight: 900 !important;
}
[data-testid="metric-container"] [data-testid="stMetricDelta"] {
    color: #8fcc6f !important;
    font-size: 13px !important;
}

/* Buttons */
.stButton > button {
    background: #ffb830 !important;
    color: #1a1008 !important;
    font-weight: 900 !important;
    font-size: 15px !important;
    border: none !important;
    border-radius: 6px !important;
    padding: 10px 20px !important;
}
.stButton > button:hover {
    background: #ffd060 !important;
    transform: translateY(-1px);
}

/* Download button */
.stDownloadButton > button {
    background: #2a4a1a !important;
    color: #8fcc6f !important;
    font-weight: 900 !important;
    font-size: 15px !important;
    border: 2px solid #8fcc6f !important;
    border-radius: 6px !important;
}

/* Info / success / warning boxes */
.stInfo, [data-testid="stAlert"][kind="info"] {
    background: #1e2e10 !important;
    border-left: 4px solid #8fcc6f !important;
    color: #c8e8a8 !important;
    font-size: 15px !important;
}
.stSuccess, [data-testid="stAlert"][kind="success"] {
    background: #1a3010 !important;
    border-left: 4px solid #5fb840 !important;
    color: #b0e890 !important;
    font-size: 15px !important;
}
.stWarning, [data-testid="stAlert"][kind="warning"] {
    background: #3a2800 !important;
    border-left: 4px solid #ffb830 !important;
    color: #ffe0a0 !important;
    font-size: 15px !important;
}
.stError, [data-testid="stAlert"][kind="error"] {
    background: #3a1010 !important;
    border-left: 4px solid #e05050 !important;
    color: #f8c0b0 !important;
    font-size: 15px !important;
}

/* Dividers */
hr { border-color: #5c3d1e !important; }

/* Spinner */
.stSpinner { color: #ffb830 !important; }

/* Radio */
.stRadio [role="radiogroup"] label { color: #f0d8b8 !important; font-size: 15px !important; }

/* File uploader */
[data-testid="stFileUploader"] {
    background: #2a1808 !important;
    border: 2px dashed #8b5e2a !important;
    border-radius: 8px !important;
    color: #f0d8b8 !important;
}

/* Number input */
[data-testid="stNumberInput"] input {
    background: #2a1808 !important;
    color: #f5e6d0 !important;
    font-size: 15px !important;
}

/* Slider */
[data-testid="stSlider"] [role="slider"] {
    background: #ffb830 !important;
}

/* Sidebar (if shown) */
section[data-testid="stSidebar"] {
    background-color: #1a1008 !important;
    border-right: 1px solid #5c3d1e;
}

/* Select slider */
.stSelectSlider { color: #f0d8b8 !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  PLOTLY THEME  — paleta más cálida
# ─────────────────────────────────────────────
PLOTLY_LAYOUT = dict(
    paper_bgcolor="#1a1008", plot_bgcolor="#221508",
    font=dict(family="Lato, Source Sans Pro, sans-serif", color="#f0d8b8", size=14),
    xaxis=dict(gridcolor="#3a2510", linecolor="#5c3d1e", tickfont=dict(size=13), color="#f0d8b8"),
    yaxis=dict(gridcolor="#3a2510", linecolor="#5c3d1e", tickfont=dict(size=13), color="#f0d8b8"),
    legend=dict(font=dict(size=13), bgcolor="#2a1808", bordercolor="#5c3d1e"),
    margin=dict(l=55, r=35, t=55, b=55),
)
COLORS = ["#ffb830", "#5fb840", "#5898f8", "#f85149"]

def render_table(df):
    rows = "".join(
        f"<tr>{''.join(f'<td>{v}</td>' for v in row)}</tr>"
        for _, row in df.iterrows()
    )
    headers = "".join(f"<th>{c}</th>" for c in df.columns)
    return f'<table class="styled-table"><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table>'

# ─────────────────────────────────────────────
#  DEFAULTS
# ─────────────────────────────────────────────
DEFAULT_ESPEC = {"Fc_kgcm2":20,"Fcr_kgcm2":30,"Slump_pulg":0,"Agua_industrial_Lm3":72,"Aire_atrapado_pct":2}
DEFAULT_AGRE  = {"Tamano_maximo_pulg":3.0,"Tamano_nominal_pulg":2.0,"Modulo_fineza":8.4,
                 "Peso_unitario_suelto_kgm3":1787.0,"Peso_unitario_compactado_kgm3":1914.0,
                 "Peso_especifico_agregado_kgm3":3012.0,"Humedad_natural_pct":3.1,
                 "Absorcion_pct":0.6,"Peso_especifico_cemento_kgm3":3110.0,"Densidad_agua_kgm3":1.0}
DEFAULT_INFRA = {"Bomba_estacionaria_m3h":50.0,"Chancadora_secundaria_th":120.0,
                 "Faja_transportadora_th":300.0,"Tolva_regulacion_m3":25.0,
                 "Personal_operadores":3,"Personal_mecanicos":1,"Personal_supervisores":1,
                 "Largo_camara_m":18.0,"Ancho_util_m":4.0,"Alto_util_m":4.5,
                 "Espesor_concreto_m":0.20,"Espesor_shotcrete_m":0.10,
                 "Espesor_relleno_estructural_m":0.40,"Altura_buzon_m":3.0,"Ancho_bandeja_m":1.2}
DEFAULT_GEO   = {"Altura_caseron_H_m":25.0,"Ancho_caseron_B_m":6.0,"Longitud_caseron_L_m":20.0,
                 "Densidad_CRF_kgm3":2200.0,"Cohesion_CRF_kPa":150.0,"Angulo_friccion_phi_deg":35.0,
                 "Densidad_roca_kgm3":2700.0,"Factor_seguridad":1.2,"Coeficiente_Poisson":0.28}

# ─────────────────────────────────────────────
#  MONTE CARLO CONSTANTS
# ─────────────────────────────────────────────
PARAM_INFO = {
    "gamma": {"label":"Peso unitario γ","unit":"MN/m³","hint":"Duro:0.028–0.032 | Medio:0.022–0.028 | Blando:0.018–0.022","default":0.024,"std_frac":0.10,"min":0.010,"max":0.040,"step":0.001},
    "H":     {"label":"Altura del relleno H","unit":"m","hint":"Bajo:<10 | Medio:10–30 | Alto:>30","default":30.0,"std_frac":0.05,"min":1.0,"max":150.0,"step":0.5},
    "B":     {"label":"Ancho expuesto B","unit":"m","hint":"Estrecho:<10 | Medio:10–20 | Amplio:>20","default":15.0,"std_frac":0.05,"min":1.0,"max":100.0,"step":0.5},
    "D":     {"label":"Longitud del bloque D","unit":"m","hint":"Corto:<10 | Medio:10–30 | Largo:>30","default":15.0,"std_frac":0.10,"min":1.0,"max":100.0,"step":0.5},
    "phi":   {"label":"Ángulo de fricción φ","unit":"grados","hint":"Bajo:<25 | Medio:25–35 | Alto:>35","default":30.0,"std_frac":0.10,"min":15.0,"max":50.0,"step":0.5},
    "L":     {"label":"Luz del techo L","unit":"m","hint":"Pequeño:<10 | Medio:10–25 | Grande:>25","default":20.0,"std_frac":0.05,"min":1.0,"max":100.0,"step":0.5},
    "t":     {"label":"Espesor del techo t","unit":"m","hint":"Delgado:<3 | Medio:3–6 | Grueso:>6","default":5.0,"std_frac":0.05,"min":0.5,"max":30.0,"step":0.1},
    "FS":    {"label":"Factor de seguridad FS","unit":"-","hint":"Bajo:1.2 | Medio:1.3–1.5 | Alto:>1.5","default":1.5,"std_frac":0.00,"min":1.0,"max":3.0,"step":0.05},
}
METHOD_COLORS = ["#ffb830","#5fb840","#5898f8","#f85149"]
METHOD_NAMES  = ["Yu (2006)","Potvin (2005)","Mitchell (1982)","Obert y Duvall (1967)"]

# ─────────────────────────────────────────────
#  GRANULOMETRÍA — ASTM D6913 SIEVES
# ─────────────────────────────────────────────
SIEVES_MM   = [75.0, 50.0, 37.5, 25.0, 19.0, 9.5, 4.75, 2.00, 0.850, 0.425, 0.250, 0.150, 0.106, 0.075]
SIEVES_NAME = ['3"','2"','1½"','1"','¾"','⅜"','N°4','N°10','N°20','N°40','N°60','N°100','N°140','N°200']

# ─────────────────────────────────────────────
#  CSV TEMPLATE
# ─────────────────────────────────────────────
def generate_template_csv() -> bytes:
    out = io.StringIO()
    out.write("# CSV MAESTRO CRF\n#\n")
    out.write("SECCION,DESIGNS\nDiseño,Cemento (kg/m³),Agua (L/m³),Desmonte (kg/m³),Densidad (kg/m³)\n")
    for r in [["D01",75,60,2812,2947],["D02",90,72,2760,2922],
              ["D03",103,82,2717,2902],["D04",138,110,2595,2843]]:
        out.write(",".join(str(v) for v in r)+"\n")
    out.write("#\nSECCION,RESISTENCIAS\nDías,D01,D02,D03,D04\n")
    for r in [[7,0.65,1.07,1.68,2.58],[14,1.01,1.37,2.73,3.14],
              [21,1.39,1.63,3.38,4.41],[28,1.66,2.08,4.82,5.21],[56,1.68,2.11,4.95,6.53]]:
        out.write(",".join(str(v) for v in r)+"\n")
    out.write("#\nSECCION,ESPECIFICACIONES\nParametro,Valor\n")
    for r in [("Fc_kgcm2",20),("Fcr_kgcm2",30),("Slump_pulg",0),("Agua_industrial_Lm3",72),("Aire_atrapado_pct",2)]:
        out.write(f"{r[0]},{r[1]}\n")
    out.write("#\nSECCION,AGREGADO\nParametro,Valor\n")
    for r in [("Tamano_maximo_pulg",3),("Tamano_nominal_pulg",2),("Modulo_fineza",8.4),
              ("Peso_unitario_suelto_kgm3",1787),("Peso_unitario_compactado_kgm3",1914),
              ("Peso_especifico_agregado_kgm3",3012),("Humedad_natural_pct",3.1),
              ("Absorcion_pct",0.6),("Peso_especifico_cemento_kgm3",3110),("Densidad_agua_kgm3",1.0)]:
        out.write(f"{r[0]},{r[1]}\n")
    out.write("#\nSECCION,INFRAESTRUCTURA\nParametro,Valor\n")
    for r in [("Bomba_estacionaria_m3h",50),("Chancadora_secundaria_th",120),
              ("Faja_transportadora_th",300),("Tolva_regulacion_m3",25),
              ("Personal_operadores",3),("Personal_mecanicos",1),("Personal_supervisores",1),
              ("Largo_camara_m",18.0),("Ancho_util_m",4.0),("Alto_util_m",4.5),
              ("Espesor_concreto_m",0.20),("Espesor_shotcrete_m",0.10),
              ("Espesor_relleno_estructural_m",0.40),("Altura_buzon_m",3.0),("Ancho_bandeja_m",1.2)]:
        out.write(f"{r[0]},{r[1]}\n")
    out.write("#\nSECCION,GEOMECANICA\nParametro,Valor\n")
    for r in [("Altura_caseron_H_m",25.0),("Ancho_caseron_B_m",6.0),("Longitud_caseron_L_m",20.0),
              ("Densidad_CRF_kgm3",2200.0),("Cohesion_CRF_kPa",150.0),("Angulo_friccion_phi_deg",35.0),
              ("Densidad_roca_kgm3",2700.0),("Factor_seguridad",1.2),("Coeficiente_Poisson",0.28)]:
        out.write(f"{r[0]},{r[1]}\n")
    # Granulometría section
    out.write("#\nSECCION,GRANULOMETRIA\nTamiz_mm,% Retenido Parcial\n")
    default_gran = [(75.0,0),(50.0,2),(37.5,5),(25.0,8),(19.0,10),(9.5,15),
                    (4.75,18),(2.00,16),(0.850,12),(0.425,8),(0.250,4),(0.150,1.5),(0.106,0.5),(0.075,0)]
    for mm, pct in default_gran:
        out.write(f"{mm},{pct}\n")
    return out.getvalue().encode("utf-8")

# ─────────────────────────────────────────────
#  CSV PARSER
# ─────────────────────────────────────────────
def parse_master_csv(uploaded_file) -> dict:
    content = uploaded_file.read().decode("utf-8")
    sections, current_section, current_rows, header = {}, None, [], None
    for line in content.splitlines():
        line = line.strip()
        if not line or line.startswith("#"): continue
        if line.upper().startswith("SECCION,"):
            if current_section and header and current_rows:
                sections[current_section] = pd.DataFrame(current_rows, columns=header)
            current_section = line.split(",",1)[1].strip().upper()
            current_rows, header = [], None
            continue
        parts = line.split(",")
        if header is None:
            header = parts
        else:
            parsed = []
            for v in parts:
                try: parsed.append(float(v) if "." in v else int(v))
                except ValueError: parsed.append(v.strip())
            current_rows.append(parsed)
    if current_section and header and current_rows:
        sections[current_section] = pd.DataFrame(current_rows, columns=header)
    return sections

# ─────────────────────────────────────────────
#  GRANULOMETRÍA — CALCULATIONS
# ─────────────────────────────────────────────
def calc_granulometria(retained_pct: list, sieves_mm: list = SIEVES_MM) -> dict:
    n = len(sieves_mm)
    arr = np.array(retained_pct, dtype=float)
    total = arr.sum()
    if total > 0:
        arr = arr / total * 100.0
    cum_ret = np.cumsum(arr)
    pct_pasa = 100.0 - cum_ret
    grava_pct   = sum(arr[i] for i, mm in enumerate(sieves_mm) if mm >= 4.75)
    arena_pct   = sum(arr[i] for i, mm in enumerate(sieves_mm) if 0.075 <= mm < 4.75)
    finos_pct   = 100.0 - grava_pct - arena_pct

    def get_D(pct_target):
        for i in range(len(pct_pasa)-1):
            if pct_pasa[i] >= pct_target >= pct_pasa[i+1]:
                p1, p2 = pct_pasa[i], pct_pasa[i+1]
                d1, d2 = sieves_mm[i], sieves_mm[i+1]
                if p1 == p2: return d1
                lod1 = math.log10(d1) if d1 > 0 else -10
                lod2 = math.log10(d2) if d2 > 0 else -10
                frac = (pct_target - p1) / (p2 - p1)
                return 10 ** (lod1 + frac * (lod2 - lod1))
        if pct_target > pct_pasa[0]: return sieves_mm[0]
        if pct_target < pct_pasa[-1]: return sieves_mm[-1]
        return None

    D10 = get_D(10)
    D30 = get_D(30)
    D60 = get_D(60)
    Cu = D60 / D10 if (D10 and D10 > 0) else None
    Cc = (D30**2) / (D10 * D60) if (D10 and D60 and D10 > 0 and D60 > 0 and D30) else None

    mf_sieves_mm = [0.150, 0.300, 0.600, 1.18, 2.36, 4.75]
    mf_sum = 0
    for mf_mm in mf_sieves_mm:
        closest = min(range(len(sieves_mm)), key=lambda i: abs(sieves_mm[i]-mf_mm))
        if abs(sieves_mm[closest] - mf_mm) < 0.5:
            mf_sum += cum_ret[closest]
        else:
            for i in range(len(sieves_mm)-1):
                if sieves_mm[i] >= mf_mm >= sieves_mm[i+1]:
                    frac = (mf_mm - sieves_mm[i]) / (sieves_mm[i+1] - sieves_mm[i])
                    mf_sum += cum_ret[i] + frac * (cum_ret[i+1] - cum_ret[i])
                    break
    for i, mm in enumerate(sieves_mm):
        if mm > 4.75:
            mf_sum += cum_ret[i]
    modulo_fineza = mf_sum / 100.0

    def classify_uscs(grava, arena, finos, Cu_val, Cc_val, D60_val):
        if finos >= 50:
            return "ML/CL (Suelo de grano fino — requiere límites de Atterberg)"
        elif grava > arena:
            if finos < 5:
                if Cu_val and Cc_val and Cu_val >= 4 and 1 <= Cc_val <= 3:
                    return "GW — Grava bien graduada"
                return "GP — Grava mal graduada"
            elif finos <= 12:
                return "GW-GM / GP-GM (símbolo dual — requiere Atterberg)"
            else:
                return "GM / GC (requiere límites de Atterberg)"
        else:
            if finos < 5:
                if Cu_val and Cc_val and Cu_val >= 6 and 1 <= Cc_val <= 3:
                    return "SW — Arena bien graduada"
                return "SP — Arena mal graduada"
            elif finos <= 12:
                return "SW-SM / SP-SM (símbolo dual — requiere Atterberg)"
            else:
                return "SM / SC (requiere límites de Atterberg)"

    clasificacion = classify_uscs(grava_pct, arena_pct, finos_pct, Cu, Cc, D60)
    tam_max_nominal = None
    for i, mm in enumerate(sieves_mm):
        if pct_pasa[i] >= 90:
            tam_max_nominal = mm
            break

    return {
        "sieves_mm": sieves_mm,
        "sieves_name": SIEVES_NAME,
        "retained_pct": arr.tolist(),
        "cum_ret": cum_ret.tolist(),
        "pct_pasa": pct_pasa.tolist(),
        "grava_pct": grava_pct,
        "arena_pct": arena_pct,
        "finos_pct": finos_pct,
        "D10": D10,
        "D30": D30,
        "D60": D60,
        "Cu": Cu,
        "Cc": Cc,
        "modulo_fineza": modulo_fineza,
        "clasificacion": clasificacion,
        "tam_max_nominal": tam_max_nominal,
    }

def default_gradacion_CRF(agre: dict) -> list:
    tam_max_pulg = agre.get("Tamano_maximo_pulg", 3.0)
    tam_nom_pulg = agre.get("Tamano_nominal_pulg", 2.0)
    mf = agre.get("Modulo_fineza", 8.4)
    tam_max_mm = tam_max_pulg * 25.4
    tam_nom_mm = tam_nom_pulg * 25.4
    retained = []
    for mm in SIEVES_MM:
        if mm > tam_max_mm:
            retained.append(0.0)
        elif mm > tam_nom_mm:
            retained.append(3.0)
        elif mm >= 19.0:
            retained.append(8.0)
        elif mm >= 9.5:
            retained.append(14.0)
        elif mm >= 4.75:
            retained.append(18.0)
        elif mm >= 2.00:
            retained.append(16.0)
        elif mm >= 0.850:
            retained.append(13.0)
        elif mm >= 0.425:
            retained.append(10.0)
        elif mm >= 0.250:
            retained.append(7.0)
        elif mm >= 0.150:
            retained.append(5.0)
        elif mm >= 0.106:
            retained.append(3.0)
        else:
            retained.append(3.0)
    return retained

# ─────────────────────────────────────────────
#  MINING PROFILE DRAW
# ─────────────────────────────────────────────
def draw_mining_profile(largo, ancho, alto, esp_conc, esp_shot, esp_rell,
                        alt_buz, anc_ban, pendiente):
    piso_drop = largo * (abs(pendiente) / 100)
    fig, ax = plt.subplots(figsize=(12, 5))
    ax.set_facecolor("#221508")
    fig.patch.set_facecolor("#1a1008")
    ax.add_patch(Rectangle((0, 0), largo, alto, edgecolor="#ffb830", facecolor="none", linewidth=2))
    ax.plot([0, largo], [0, -piso_drop], color="#5fb840", linewidth=2)
    hatch_color = "#5c3d1e"
    hatch_spacing = 0.4
    x_min, x_max = 0, largo
    y_min, y_max = -piso_drop - 3, alto + esp_shot + 3
    for i in np.arange(-50, largo + 50, hatch_spacing):
        ax.plot([i, i + 50], [y_min, y_max], color=hatch_color, linewidth=0.4, alpha=0.35)
    ax.add_patch(Rectangle((0, 0), largo, alto, facecolor="#221508", edgecolor="none"))
    ax.add_patch(Rectangle((0, alto), largo, esp_shot, edgecolor="#5898f8", facecolor="none",
                            linestyle="--", linewidth=1.8))
    lhd_x = largo * 0.05
    lhd_y = -0.2
    ax.add_patch(Rectangle((lhd_x, lhd_y), 2.5, 0.8, facecolor="#e3b341", edgecolor="black", linewidth=1))
    ax.add_patch(Rectangle((lhd_x + 1.8, lhd_y + 0.4), 0.7, 0.6, facecolor="#d4a12c", edgecolor="black", linewidth=1))
    ax.add_patch(Polygon([(lhd_x-0.2,lhd_y),(lhd_x+0.3,lhd_y-0.4),(lhd_x+1.0,lhd_y-0.4),(lhd_x+1.0,lhd_y)],
                          closed=True, facecolor="#e3b341", edgecolor="black"))
    for w in [0.3, 1.9]:
        ax.add_patch(plt.Circle((lhd_x + w, lhd_y), 0.25, facecolor="#1a1008", edgecolor="white", linewidth=1))
    ax.set_title("Perfil Técnico — Cámara de Mezclado CRF", color="#ffb830", fontsize=13)
    ax.set_xlabel("Longitud (m)", color="#f0d8b8")
    ax.set_ylabel("Altura (m)", color="#f0d8b8")
    ax.tick_params(colors="#f0d8b8")
    ax.grid(color="#3a2510", linestyle="--", linewidth=0.5)
    ax.set_xlim(-2, largo + 2)
    ax.set_ylim(-piso_drop - 2, alto + esp_shot + 3)
    return fig

# ─────────────────────────────────────────────
#  MONTE CARLO FUNCTIONS
# ─────────────────────────────────────────────
def calc_alpha(phi):        return 45 + phi / 2
def calc_ucs_yu(gamma, H): return 2.5 * gamma * H

def calc_ucs_potvin(gamma, B, H, D):
    return (gamma * B * (2 * H - D)) / (2 * H + 2 * B - D)

def calc_ucs_mitchell(gamma, B, H, D, phi):
    alpha = np.radians(calc_alpha(phi))
    Hs    = H - (D * np.tan(alpha) / 2)
    num   = gamma * Hs * B
    den   = Hs + B * np.tan(alpha)
    c     = num / (2 * den)
    return (2 * c * np.cos(np.radians(phi))) / (1 - np.sin(np.radians(phi)))

def calc_ucs_roof(gamma, L, t, FS):
    return (3 * gamma * (L ** 2) / (4 * (t ** 2))) * FS

def interpret_ucs(val):
    if   val < 0.5: return "Favorable", "#7ab87a"
    elif val < 1.5: return "Moderado",  "#d4c44a"
    elif val < 3.0: return "Severo",    "#e08040"
    else:           return "Crítico",   "#d05040"

def sample_param(mean, std_frac, size):
    if std_frac == 0: return np.full(size, mean)
    std  = mean * std_frac
    a, b = (0 - mean) / std, (10 * mean - mean) / std
    return truncnorm.rvs(a, b, loc=mean, scale=std, size=size)

def monte_carlo(p, n_mc):
    keys = list(PARAM_INFO.keys())
    samples = {k: sample_param(p[k], PARAM_INFO[k]["std_frac"], n_mc) for k in keys}
    g, H, B, D, phi, L, t, FS = (samples[k] for k in keys)
    u1 = calc_ucs_yu(g, H)
    u2 = calc_ucs_potvin(g, B, H, D)
    u3 = calc_ucs_mitchell(g, B, H, D, phi)
    u4 = calc_ucs_roof(g, L, t, FS)
    inp = np.column_stack([samples[k] for k in keys])
    out = np.column_stack([u1, u2, u3, u4])
    return inp, out, keys

def mc_stats(arr):
    return dict(mean=np.mean(arr), std=np.std(arr),
                p5=np.percentile(arr,5), p50=np.percentile(arr,50),
                p90=np.percentile(arr,90), p95=np.percentile(arr,95))

# MC Plot helpers — updated with warm palette
def fig_mc_histograms(outputs):
    fig, axes = plt.subplots(2, 2, figsize=(11, 7), tight_layout=True)
    fig.patch.set_facecolor("#1a1008")
    for i, ax in enumerate(axes.flat):
        ax.set_facecolor("#221508")
        color = METHOD_COLORS[i]
        ax.hist(outputs[:, i], bins=60, density=True, color=color, alpha=.85, edgecolor="#1a1008", linewidth=.4)
        ax.axvline(np.mean(outputs[:, i]), color="#fff8e8", lw=1.8, ls="--", label=f"Media: {np.mean(outputs[:,i]):.3f}")
        ax.axvline(np.percentile(outputs[:, i], 5),  color="#c9a87c", lw=1.2, ls=":", label="P5")
        ax.axvline(np.percentile(outputs[:, i], 95), color="#c9a87c", lw=1.2, ls=":", label="P95")
        ax.set_title(METHOD_NAMES[i], fontsize=11, color=color, pad=8)
        ax.set_xlabel("UCS (MPa)", fontsize=9, color="#f0d8b8")
        ax.set_ylabel("Densidad", fontsize=9, color="#f0d8b8")
        ax.tick_params(colors="#c9a87c")
        ax.legend(fontsize=7, framealpha=.3)
        ax.grid(True, alpha=.4, color="#3a2510")
    return fig

def fig_mc_cdf(outputs):
    fig, axes = plt.subplots(2, 2, figsize=(11, 7), tight_layout=True)
    fig.patch.set_facecolor("#1a1008")
    for i, ax in enumerate(axes.flat):
        ax.set_facecolor("#221508")
        color = METHOD_COLORS[i]
        s   = np.sort(outputs[:, i])
        cdf = np.arange(len(s)) / len(s)
        ax.plot(s, cdf, color=color, lw=2)
        ax.fill_between(s, cdf, alpha=.15, color=color)
        for pct, ls in [(5,"--"),(50,"-."),(95,":")]:
            v = np.percentile(outputs[:, i], pct)
            ax.axvline(v, color="#fff8e8", lw=1, ls=ls, label=f"P{pct}: {v:.3f}")
        ax.set_title(METHOD_NAMES[i], fontsize=11, color=color, pad=8)
        ax.set_xlabel("UCS (MPa)", fontsize=9, color="#f0d8b8")
        ax.set_ylabel("Probabilidad acumulada", fontsize=9, color="#f0d8b8")
        ax.tick_params(colors="#c9a87c")
        ax.legend(fontsize=7, framealpha=.3)
        ax.grid(True, alpha=.4, color="#3a2510")
    return fig

def fig_mc_tornado(inputs, outputs, keys):
    labels = [PARAM_INFO[k]["label"] for k in keys]
    corrs  = [np.corrcoef(inputs[:, i], outputs[:, 0])[0, 1] for i in range(inputs.shape[1])]
    order  = np.argsort(np.abs(corrs))
    fig, ax = plt.subplots(figsize=(9, 5))
    fig.patch.set_facecolor("#1a1008")
    ax.set_facecolor("#221508")
    ax.barh(range(len(corrs)), [corrs[i] for i in order],
            color=["#ffb830" if corrs[i] >= 0 else "#f85149" for i in order],
            edgecolor="#1a1008", height=0.65)
    ax.set_yticks(range(len(corrs)))
    ax.set_yticklabels([labels[i] for i in order], fontsize=9, color="#f0d8b8")
    ax.axvline(0, color="#fff8e8", lw=1, alpha=.5)
    ax.set_xlabel("Correlación de Pearson con UCS Yu", fontsize=9, color="#f0d8b8")
    ax.set_title("Diagrama de Tornado — Sensibilidad de parámetros", fontsize=11, color="#ffb830")
    ax.tick_params(colors="#c9a87c")
    ax.grid(True, axis="x", alpha=.35, color="#3a2510")
    return fig, corrs

def fig_mc_scatter(inputs, outputs, keys):
    fig, axes = plt.subplots(2, 4, figsize=(14, 6), tight_layout=True)
    fig.patch.set_facecolor("#1a1008")
    for i, ax in enumerate(axes.flat):
        ax.set_facecolor("#221508")
        ax.scatter(inputs[:, i], outputs[:, 0], s=2, alpha=.25, color="#ffb830")
        ax.set_xlabel(PARAM_INFO[keys[i]]["label"], fontsize=7, color="#f0d8b8")
        ax.set_ylabel("UCS Yu (MPa)", fontsize=7, color="#f0d8b8")
        ax.set_title(PARAM_INFO[keys[i]]["label"], fontsize=8, color="#d4935a")
        ax.tick_params(colors="#c9a87c")
        ax.grid(True, alpha=.3, color="#3a2510")
    return fig

def fig_mc_box(outputs):
    fig, ax = plt.subplots(figsize=(8, 4))
    fig.patch.set_facecolor("#1a1008")
    ax.set_facecolor("#221508")
    bp = ax.boxplot([outputs[:, i] for i in range(4)], patch_artist=True,
                    medianprops=dict(color="#fff8e8", lw=2),
                    whiskerprops=dict(color="#c9a87c"),
                    capprops=dict(color="#c9a87c"),
                    flierprops=dict(marker="o", markerfacecolor="#c9a87c", markersize=2, alpha=.3))
    for patch, color in zip(bp["boxes"], METHOD_COLORS):
        patch.set_facecolor(color); patch.set_alpha(.75)
    ax.set_xticks(range(1, 5))
    ax.set_xticklabels(METHOD_NAMES, color="#f0d8b8")
    ax.set_ylabel("UCS (MPa)", color="#f0d8b8")
    ax.set_title("Distribución de UCS por método", fontsize=11, color="#ffb830")
    ax.tick_params(colors="#c9a87c")
    ax.grid(True, axis="y", alpha=.35, color="#3a2510")
    return fig

# ─────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def export_mc_excel(p, det, mc_stats_list, inputs, outputs, corr, keys):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "1_Datos_Entrada"
    ws1.append(["Parámetro","Valor","Unidad","Típicos","CoV (%)"])
    for cell in ws1[1]: cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    for k, v in PARAM_INFO.items():
        ws1.append([v["label"], p[k], v["unit"], v["hint"], v["std_frac"]*100])
    for col in ["A","B","C","D","E"]: ws1.column_dimensions[col].width = 28
    ws2 = wb.create_sheet("2_Interpretacion")
    ws2.append(["Método","UCS (MPa)","Clasificación"])
    for cell in ws2[1]: cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    COLOR_MAP = {"Favorable":"00CC44","Moderado":"DDDD00","Severo":"FF8800","Crítico":"FF2200"}
    for k, val in det.items():
        txt, _ = interpret_ucs(val)
        ws2.append([k, round(val,4), txt])
        ws2.cell(ws2.max_row,3).fill = PatternFill(start_color=COLOR_MAP[txt], end_color=COLOR_MAP[txt], fill_type="solid")
    ws3 = wb.create_sheet("3_MonteCarlo_Estadisticas")
    ws3.append(["Método","Media","Std","P5","P50","P90","P95"])
    for cell in ws3[1]: cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    for i, name in enumerate(METHOD_NAMES):
        s = mc_stats_list[i]
        ws3.append([name, round(s["mean"],4), round(s["std"],4), round(s["p5"],4),
                    round(s["p50"],4), round(s["p90"],4), round(s["p95"],4)])
    ws4 = wb.create_sheet("4_Sensibilidad")
    ws4.append(["Parámetro","Correlación con UCS Yu"])
    for cell in ws4[1]: cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    for k, c in zip(keys, corr):
        ws4.append([PARAM_INFO[k]["label"], round(c,4)])
    ws5 = wb.create_sheet("5_Muestra_MC")
    header = [PARAM_INFO[k]["label"] for k in keys] + METHOD_NAMES
    ws5.append(header)
    for cell in ws5[1]: cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    for i in range(min(1000, len(outputs))):
        ws5.append(list(inputs[i]) + list(outputs[i]))
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.getvalue()

# ─────────────────────────────────────────────
#  PDF BUILDER
# ─────────────────────────────────────────────
_GOLD  = colors.HexColor("#ffb830")
_DARK2 = colors.HexColor("#3a2005")
_ROW1  = colors.HexColor("#2a1808")
_ROW2  = colors.HexColor("#1a1008")
_LIGHT = colors.HexColor("#f0d8b8")

def _tbl_style():
    return TableStyle([
        ("BACKGROUND",    (0,0),(-1,0),  _DARK2),
        ("TEXTCOLOR",     (0,0),(-1,0),  _GOLD),
        ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,-1), 9),
        ("ALIGN",         (0,0),(-1,-1), "LEFT"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1), [_ROW1, _ROW2]),
        ("TEXTCOLOR",     (0,1),(-1,-1), _LIGHT),
        ("GRID",          (0,0),(-1,-1), 0.3, colors.HexColor("#5c3d1e")),
        ("LEFTPADDING",   (0,0),(-1,-1), 6),
        ("RIGHTPADDING",  (0,0),(-1,-1), 6),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
    ])

def _side_style():
    return TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LEFTPADDING",(0,0),(-1,-1),0),
        ("RIGHTPADDING",(0,0),(-1,-1),0),
        ("TOPPADDING",(0,0),(-1,-1),0),
        ("BOTTOMPADDING",(0,0),(-1,-1),0),
    ])

def df2tbl(df, widths):
    data = [list(df.columns)] + [list(r) for _,r in df.iterrows()]
    t = Table(data, colWidths=widths, repeatRows=1)
    t.setStyle(_tbl_style())
    return t

def side_by_side(left, right, lw=8.5*cm, rw=8.5*cm, gap=0.5*cm):
    t = Table([[left, Spacer(gap,1), right]], colWidths=[lw, gap, rw])
    t.setStyle(_side_style())
    return t

def plotly_img(fig, w=15*cm, h=7*cm):
    buf = io.BytesIO()
    fig.write_image(buf, format="png", width=950, height=440, scale=2)
    buf.seek(0)
    return RLImage(buf, width=w, height=h)

def mpl_img(fig, w=15*cm, h=6*cm):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=160, bbox_inches="tight", facecolor=fig.get_facecolor())
    buf.seek(0)
    return RLImage(buf, width=w, height=h)

def build_pdf_report(s: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    base = getSampleStyleSheet()
    base.add(ParagraphStyle("T1", parent=base["Normal"], fontSize=16, textColor=_GOLD,
        fontName="Helvetica-Bold", spaceAfter=4, alignment=TA_CENTER))
    base.add(ParagraphStyle("T2", parent=base["Normal"], fontSize=11, textColor=_GOLD,
        fontName="Helvetica-Bold", spaceBefore=8, spaceAfter=3))
    base.add(ParagraphStyle("T3", parent=base["Normal"], fontSize=9.5, textColor=_GOLD,
        fontName="Helvetica-Bold", spaceBefore=5, spaceAfter=2))
    base.add(ParagraphStyle("SB", parent=base["Normal"], fontSize=8.5, textColor=_LIGHT, spaceAfter=4))
    base.add(ParagraphStyle("SB2", parent=base["Normal"], fontSize=8, textColor=_LIGHT, spaceAfter=2))

    def sec(txt): return Paragraph(txt, base["T2"])
    def sub(txt): return Paragraph(txt, base["T3"])
    def sp(n=0.3): return Spacer(1, n*cm)
    def hr(): return HRFlowable(width="100%", thickness=0.5, color=_DARK2, spaceAfter=4)

    hoy      = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
    selected = s["selected"]
    story    = []

    story += [sp(1),
        Paragraph("REPORTE TECNICO — CEMENTED ROCK FILL (CRF)", base["T1"]),
        Paragraph("Metodo Sub Level Stoping | Análisis Completo CRF + Granulometría + Monte Carlo", base["SB"]),
        HRFlowable(width="100%", thickness=1.5, color=_GOLD, spaceAfter=6),
        Paragraph(f"Generado: {hoy}   |   Diseño activo: <b>{selected}</b>", base["SB"]),
        sp(0.5)]

    story += [sec("1. DISEÑOS DE MEZCLA"), hr()]
    df_mix = s["designs_df"].copy()
    if "A/C" not in df_mix.columns:
        df_mix["A/C"] = (df_mix["Agua (L/m³)"]/df_mix["Cemento (kg/m³)"]).round(3)
    show = df_mix[["Diseño","Cemento (kg/m³)","Agua (L/m³)","Desmonte (kg/m³)","Densidad (kg/m³)","A/C"]]
    story.append(df2tbl(show, [2*cm,3*cm,2.5*cm,3*cm,3*cm,2*cm]))
    story += [sp(0.25), sub("Resistencias a Compresion (MPa)")]
    story.append(df2tbl(s["resist_df"], [2*cm,3.5*cm,3.5*cm,3.5*cm,3.5*cm]))
    story += [sp(0.25)]
    resist_df = s["resist_df"]
    fig_res = go.Figure()
    for d, c in zip(["D01","D02","D03","D04"],["#ffb830","#5fb840","#5898f8","#f85149"]):
        fig_res.add_trace(go.Scatter(x=resist_df["Días"], y=resist_df[d], mode="lines+markers",
            name=d, opacity=1.0 if d==selected else 0.4,
            line=dict(color=c, width=3 if d==selected else 1.5)))
    fig_res.update_layout(**PLOTLY_LAYOUT, height=320, title=f"Curva de Resistencia — {selected}")
    story.append(plotly_img(fig_res, w=15*cm, h=6.5*cm))

    story += [PageBreak(), sec("2. DISEÑO SELECCIONADO"), hr()]
    espec=s["espec"]; agre=s["agre"]
    row_d = df_mix[df_mix["Diseño"]==selected].iloc[0]
    cem=float(row_d["Cemento (kg/m³)"]); desm=float(row_d["Desmonte (kg/m³)"])
    rel_ac=float(row_d["A/C"]); agua_ind=float(espec["Agua_industrial_Lm3"])
    aire=int(espec["Aire_atrapado_pct"]); hum=float(agre["Humedad_natural_pct"])
    absorc=float(agre["Absorcion_pct"])
    agg_corr = desm/(1+hum/100)
    aporte   = agg_corr*((hum-absorc)/100)
    dens_m   = cem+agua_ind+aporte+agg_corr
    df_esp = pd.DataFrame({"Parametro":["F'c (kg/cm2)","F'cr (kg/cm2)","Slump (pulg)","Agua ind. (L/m3)","Aire (%)","A/C"],
        "Valor":[f"{espec['Fc_kgcm2']:.0f}",f"{espec['Fcr_kgcm2']:.0f}",f"{espec['Slump_pulg']:.1f}",
                 f"{agua_ind:.1f}",f"{aire}",f"{rel_ac:.3f}"]})
    df_agg = pd.DataFrame({"Parametro":["Tam. max.","Tam. nom.","Mod. fineza",
                                        "P.U. suelto","P.U. compact.","P.E. agr.","Hum. nat. (%)","Absorcion (%)"],
        "Valor":[f"{agre['Tamano_maximo_pulg']:.2f} pulg",f"{agre['Tamano_nominal_pulg']:.2f} pulg",
                 f"{agre['Modulo_fineza']:.2f}",f"{agre['Peso_unitario_suelto_kgm3']:.0f} kg/m3",
                 f"{agre['Peso_unitario_compactado_kgm3']:.0f} kg/m3",f"{agre['Peso_especifico_agregado_kgm3']:.0f} kg/m3",
                 f"{agre['Humedad_natural_pct']:.2f}",f"{agre['Absorcion_pct']:.2f}"]})
    story.append(side_by_side(df2tbl(df_esp,[5*cm,3*cm]), df2tbl(df_agg,[5.5*cm,3.5*cm])))
    df_calc = pd.DataFrame({"Concepto":["Cemento (kg/m3)","Agua ind. (L/m3)","Agg. corregido (kg/m3)",
                                        "Aporte agua (L/m3)","Aire atrapado (%)","Densidad mezcla (kg/m3)"],
        "Valor":[f"{cem:.2f}",f"{agua_ind:.2f}",f"{agg_corr:.2f}",f"{aporte:.2f}",f"{aire}%",f"{dens_m:.2f}"]})
    story += [sp(0.25), sub("Calculo del Diseño de Mezcla")]
    story.append(df2tbl(df_calc,[9*cm,6*cm]))

    story += [PageBreak(), sec("3. INFRAESTRUCTURA — CAMARA DE MEZCLADO"), hr()]
    infra=s["infra"]; pend=s["pendiente"]
    largo=float(infra["Largo_camara_m"]); ancho=float(infra["Ancho_util_m"]); alto=float(infra["Alto_util_m"])
    esp_conc=float(infra["Espesor_concreto_m"]); esp_shot=float(infra["Espesor_shotcrete_m"])
    esp_rell=float(infra["Espesor_relleno_estructural_m"])
    alt_buz=float(infra["Altura_buzon_m"]); anc_ban=float(infra["Ancho_bandeja_m"])
    piso_drop=largo*(abs(pend)/100)
    vol_util=largo*ancho*alto; vol_conc=largo*ancho*esp_conc
    vol_exc=(ancho+2*esp_conc)*(alto+2*esp_conc)*largo
    df_eq = pd.DataFrame({"Equipo":["Bomba estacionaria","Chancadora secundaria","Faja transportadora","Tolva de regulacion"],
        "Capacidad":[f"{infra['Bomba_estacionaria_m3h']:.0f} m3/h",f"{infra['Chancadora_secundaria_th']:.0f} t/h",
                     f"{infra['Faja_transportadora_th']:.0f} t/h",f"{infra['Tolva_regulacion_m3']:.0f} m3"]})
    df_vol = pd.DataFrame({"Concepto":["Volumen util (m3)","Caida piso (m)","Vol. concreto (m3)","Vol. excavado (m3)"],
        "Valor":[f"{vol_util:.2f}",f"{piso_drop:.2f}",f"{vol_conc:.2f}",f"{vol_exc:.2f}"]})
    story.append(side_by_side(df2tbl(df_eq,[5.5*cm,3.5*cm]), df2tbl(df_vol,[5.5*cm,3*cm])))
    story += [sp(0.3)]
    fig_tec = draw_mining_profile(largo, ancho, alto, esp_conc, esp_shot, esp_rell, alt_buz, anc_ban, pend)
    story.append(mpl_img(fig_tec, w=16*cm, h=7.5*cm))
    plt.close(fig_tec)

    story += [PageBreak(), sec("4. CALCULOS GEOMECANICOS — MITCHELL + DUJISIN"), hr()]
    geo=s["geo"]
    H=float(geo["Altura_caseron_H_m"]); dens_rell=float(geo["Densidad_CRF_kgm3"])
    phi=float(geo["Angulo_friccion_phi_deg"]); fs=float(geo["Factor_seguridad"])
    dens_roca=float(geo["Densidad_roca_kgm3"])
    gamma_rell=dens_rell*9.81/1e6; gamma_roca=dens_roca*9.81/1e6
    K=(1-math.sin(math.radians(phi)))/(1+math.sin(math.radians(phi)))
    sigma_v=gamma_rell*H; sigma_h=K*sigma_v
    sigma_t=sigma_v/fs; sigma_l=sigma_h/fs; sigma_tot=sigma_t+sigma_l
    df_p1 = pd.DataFrame({"Parametro":["H (m)","B (m)","L (m)","Dens. CRF (kg/m3)","Cohesion (kPa)","Phi (deg)"],
        "Valor":[f"{geo['Altura_caseron_H_m']:.1f}",f"{geo['Ancho_caseron_B_m']:.1f}",f"{geo['Longitud_caseron_L_m']:.1f}",
                 f"{dens_rell:.0f}",f"{geo['Cohesion_CRF_kPa']:.0f}",f"{phi:.1f}"]})
    df_p2 = pd.DataFrame({"Parametro":["Dens. roca (kg/m3)","Factor seg.","Poisson nu","g CRF (MN/m3)"],
        "Valor":[f"{dens_roca:.0f}",f"{fs}",f"{geo['Coeficiente_Poisson']:.2f}",f"{gamma_rell:.6f}"]})
    story.append(side_by_side(df2tbl(df_p1,[5*cm,3*cm]), df2tbl(df_p2,[5*cm,3*cm])))
    df_mit = pd.DataFrame({"Concepto":["sv (MN/m2)","K activo","sh (MN/m2)","s total (MN/m2)"],
        "Valor":[f"{sigma_v:.4f}",f"{K:.4f}",f"{sigma_h:.4f}",f"{sigma_tot:.4f}"]})
    story += [sp(0.25), df2tbl(df_mit, [9*cm, 6*cm])]
    z = np.linspace(0, H, 50)
    fig_geo = go.Figure()
    fig_geo.add_trace(go.Scatter(x=K*gamma_rell*z, y=z, name="Mitchell", mode="lines", line=dict(color="#ffb830",width=3)))
    fig_geo.add_trace(go.Scatter(x=K*gamma_rell*z/fs, y=z, name="Dujisin", mode="lines", line=dict(color="#5fb840",width=3,dash="dash")))
    fig_geo.update_layout(**PLOTLY_LAYOUT, height=340, title="Distribucion de Presion Lateral (MN/m2)",
        xaxis_title="Presion (MN/m2)", yaxis_title="Altura (m)")
    story.append(plotly_img(fig_geo, w=15*cm, h=7*cm))

    if s.get("gran_data"):
        story += [PageBreak(), sec("5. ANÁLISIS GRANULOMÉTRICO — ASTM D6913 / D2487 / D1140"), hr()]
        gd = s["gran_data"]
        df_gran = pd.DataFrame({
            "Tamiz": gd["sieves_name"],
            "Abertura (mm)": [f"{mm:.3f}" for mm in gd["sieves_mm"]],
            "% Ret. Parcial": [f"{v:.2f}" for v in gd["retained_pct"]],
            "% Ret. Acum.": [f"{v:.2f}" for v in gd["cum_ret"]],
            "% Que Pasa": [f"{v:.2f}" for v in gd["pct_pasa"]],
        })
        story.append(df2tbl(df_gran, [1.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]))
        story += [sp(0.3)]
        d10_str = f"{gd['D10']:.4f} mm" if gd['D10'] else "N/A"
        d30_str = f"{gd['D30']:.4f} mm" if gd['D30'] else "N/A"
        d60_str = f"{gd['D60']:.4f} mm" if gd['D60'] else "N/A"
        cu_str  = f"{gd['Cu']:.2f}" if gd['Cu'] else "N/A"
        cc_str  = f"{gd['Cc']:.3f}" if gd['Cc'] else "N/A"
        df_params = pd.DataFrame({
            "Parámetro": ["Grava (≥4.75mm)","Arena (0.075–4.75mm)","Finos (<0.075mm)",
                          "D10","D30","D60","Cu (Coef. Uniformidad)","Cc (Coef. Curvatura)",
                          "Módulo de Fineza","Clasificación USCS (D2487)"],
            "Valor": [f"{gd['grava_pct']:.1f}%", f"{gd['arena_pct']:.1f}%", f"{gd['finos_pct']:.1f}%",
                      d10_str, d30_str, d60_str, cu_str, cc_str,
                      f"{gd['modulo_fineza']:.2f}", gd['clasificacion']]
        })
        story.append(df2tbl(df_params, [7*cm, 8.5*cm]))
        fig_gran_pdf, ax_g = plt.subplots(figsize=(12, 5))
        fig_gran_pdf.patch.set_facecolor("#1a1008")
        ax_g.set_facecolor("#221508")
        ax_g.semilogx(gd["sieves_mm"], gd["pct_pasa"], "o-", color="#ffb830", linewidth=2.5,
                      markersize=6, markerfacecolor="#ffb830")
        ax_g.fill_between(gd["sieves_mm"], gd["pct_pasa"], alpha=0.15, color="#ffb830")
        for d_val, d_label, col in [(gd["D10"],"D₁₀","#5fb840"),(gd["D30"],"D₃₀","#5898f8"),(gd["D60"],"D₆₀","#f85149")]:
            if d_val:
                ax_g.axvline(d_val, color=col, linestyle="--", alpha=0.8, linewidth=1.5)
                ax_g.text(d_val*1.05, 5, d_label, color=col, fontsize=9)
        ax_g.axvline(4.75, color="#8b7355", linestyle=":", alpha=0.5, linewidth=1)
        ax_g.axvline(0.075, color="#8b7355", linestyle=":", alpha=0.5, linewidth=1)
        ax_g.text(4.75*1.05, 50, "N°4", color="#8b7355", fontsize=8)
        ax_g.text(0.075*1.05, 50, "N°200", color="#8b7355", fontsize=8)
        ax_g.set_xlabel("Tamaño de partícula (mm)", color="#f0d8b8", fontsize=11)
        ax_g.set_ylabel("% Que Pasa", color="#f0d8b8", fontsize=11)
        ax_g.set_title("Curva Granulométrica — ASTM D6913", color="#ffb830", fontsize=13)
        ax_g.set_ylim(0, 105); ax_g.set_xlim(min(gd["sieves_mm"])*0.5, max(gd["sieves_mm"])*2)
        ax_g.tick_params(colors="#c9a87c")
        ax_g.grid(True, which="both", color="#3a2510", linestyle="--", linewidth=0.5, alpha=0.7)
        story.append(mpl_img(fig_gran_pdf, w=16*cm, h=7*cm))
        plt.close(fig_gran_pdf)

    if s.get("mc_results"):
        story += [PageBreak(), sec("6. ANÁLISIS UCS — SIMULACIÓN MONTE CARLO"), hr()]
        mc = s["mc_results"]
        det = mc["det"]; mc_st = mc["mc_stats"]
        story += [sub("Resultados Determinísticos")]
        df_det = pd.DataFrame({
            "Método": list(det.keys()),
            "UCS (MPa)": [f"{v:.4f}" for v in det.values()],
            "Clasificación": [interpret_ucs(v)[0] for v in det.values()],
        })
        story.append(df2tbl(df_det, [5*cm, 3.5*cm, 6*cm]))
        story += [sp(0.3), sub("Estadísticas Monte Carlo")]
        rows_mc = []
        for i, name in enumerate(METHOD_NAMES):
            st2 = mc_st[i]
            rows_mc.append([name, f"{st2['mean']:.4f}", f"{st2['std']:.4f}",
                            f"{st2['p5']:.4f}", f"{st2['p50']:.4f}", f"{st2['p90']:.4f}", f"{st2['p95']:.4f}"])
        df_mc_tbl = pd.DataFrame(rows_mc, columns=["Método","Media","Std","P5","P50","P90","P95"])
        story.append(df2tbl(df_mc_tbl, [4.5*cm,1.8*cm,1.5*cm,1.8*cm,1.8*cm,1.8*cm,1.8*cm]))
        story += [sp(0.3), sub("Distribuciones Monte Carlo")]
        fig_h = fig_mc_histograms(mc["outputs"])
        story.append(mpl_img(fig_h, w=16*cm, h=8*cm))
        plt.close(fig_h)
        story += [sp(0.2), sub("Análisis de Sensibilidad — Diagrama de Tornado")]
        fig_t, _ = fig_mc_tornado(mc["inputs"], mc["outputs"], mc["keys"])
        story.append(mpl_img(fig_t, w=15*cm, h=6*cm))
        plt.close(fig_t)

    story += [sp(0.5),
        HRFlowable(width="100%", thickness=1, color=_GOLD, spaceAfter=4),
        Paragraph(f"Cemented Rock Fill · Sistema CRF  |  Desarrollado por: <b>Bradoc Chambilla</b>  |  {hoy}", base["SB"])]

    doc.build(story)
    buf.seek(0)
    return buf.read()

# ═══════════════════════════════════════════════════════════
#  BANNER
# ═══════════════════════════════════════════════════════════
st.markdown("""
<div class="hero-banner">
  <h1>⛏ Cemented Rock Fill (CRF)</h1>
  <p>Análisis técnico-económico · Sub Level Stoping · Granulometría ASTM · Simulación Monte Carlo UCS</p>
</div>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════
#  SESSION STATE
# ═══════════════════════════════════════════════════════════
if "csv_activo"  not in st.session_state: st.session_state.csv_activo  = False
if "pendiente"   not in st.session_state: st.session_state.pendiente   = -1
if "designs_df"  not in st.session_state:
    st.session_state.designs_df = pd.DataFrame({
        "Diseño":["D01","D02","D03","D04"],
        "Cemento (kg/m³)":[75.,90.,103.,138.], "Agua (L/m³)":[60.,72.,82.,110.],
        "Desmonte (kg/m³)":[2812.,2760.,2717.,2595.], "Densidad (kg/m³)":[2947.,2922.,2902.,2843.],
    })
if "resist_df" not in st.session_state:
    st.session_state.resist_df = pd.DataFrame({
        "Días":[7,14,21,28,56],
        "D01":[0.65,1.01,1.39,1.66,1.68],"D02":[1.07,1.37,1.63,2.08,2.11],
        "D03":[1.68,2.73,3.38,4.82,4.95],"D04":[2.58,3.14,4.41,5.21,6.53],
    })
if "espec"         not in st.session_state: st.session_state.espec = DEFAULT_ESPEC.copy()
if "agre"          not in st.session_state: st.session_state.agre  = DEFAULT_AGRE.copy()
if "infra"         not in st.session_state: st.session_state.infra = DEFAULT_INFRA.copy()
if "geo"           not in st.session_state: st.session_state.geo   = DEFAULT_GEO.copy()
if "gran_retained" not in st.session_state: st.session_state.gran_retained = None
if "mc_results"    not in st.session_state: st.session_state.mc_results    = None
if "mc_params"     not in st.session_state: st.session_state.mc_params     = {k: v["default"] for k,v in PARAM_INFO.items()}

# ═══════════════════════════════════════════════════════════
#  CARGA CSV
# ═══════════════════════════════════════════════════════════
st.markdown("### 📂 Carga de Datos Maestros")
col_info, col_dl = st.columns([3,1])
with col_info:
    st.info("**Paso 1:** Descargue la plantilla CSV → **Paso 2:** Complete sus datos → **Paso 3:** Suba el archivo aquí.")
with col_dl:
    st.download_button("⬇️ Descargar Plantilla CSV", data=generate_template_csv(),
        file_name="plantilla_CRF_maestro.csv", mime="text/csv", use_container_width=True)

uploaded_csv = st.file_uploader("Seleccionar archivo CSV maestro (.csv)", type=["csv"], key="master_csv")
if uploaded_csv is not None:
    try:
        sections = parse_master_csv(uploaded_csv)
        def load_kv(sec, default):
            if sec not in sections: return default.copy()
            df_ = sections[sec].copy(); df_.columns = ["Parametro","Valor"]
            d = default.copy()
            for _, r in df_.iterrows():
                k = str(r["Parametro"]).strip()
                if k in d: d[k] = float(r["Valor"])
            return d
        if "DESIGNS" in sections:
            dc = sections["DESIGNS"].copy()
            dc.columns = ["Diseño","Cemento (kg/m³)","Agua (L/m³)","Desmonte (kg/m³)","Densidad (kg/m³)"]
            for col in dc.columns[1:]: dc[col] = pd.to_numeric(dc[col], errors="coerce")
            st.session_state.designs_df = dc.reset_index(drop=True)
        if "RESISTENCIAS" in sections:
            dr = sections["RESISTENCIAS"].copy(); dr.columns = ["Días","D01","D02","D03","D04"]
            for col in dr.columns: dr[col] = pd.to_numeric(dr[col], errors="coerce")
            st.session_state.resist_df = dr.reset_index(drop=True)
        st.session_state.espec = load_kv("ESPECIFICACIONES", DEFAULT_ESPEC)
        st.session_state.agre  = load_kv("AGREGADO",        DEFAULT_AGRE)
        st.session_state.infra = load_kv("INFRAESTRUCTURA", DEFAULT_INFRA)
        st.session_state.geo   = load_kv("GEOMECANICA",     DEFAULT_GEO)
        if "GRANULOMETRIA" in sections:
            df_g = sections["GRANULOMETRIA"].copy()
            df_g.columns = ["Tamiz_mm","Retenido_pct"]
            st.session_state.gran_retained = df_g["Retenido_pct"].tolist()
        st.session_state.csv_activo = True
        st.success(f"✔ CSV cargado — secciones: {', '.join(sections.keys())}.")
    except Exception as e:
        st.error(f"❌ Error al procesar el CSV: {e}")

if st.session_state.csv_activo:
    if st.button("🔄 Restaurar datos por defecto"):
        for k in ["csv_activo","designs_df","resist_df","espec","agre","infra","geo","gran_retained"]:
            if k in st.session_state: del st.session_state[k]
        st.rerun()

st.markdown("---")

# ═══════════════════════════════════════════════════════════
#  TABS — NUEVO ORDEN: Monte Carlo ANTES de Granulometría
# ═══════════════════════════════════════════════════════════
tabs = st.tabs([
    "🧪 Diseños de Mezcla",
    "📐 Diseño Seleccionado",
    "🏗️ Infraestructura",
    "📊 Cálc. Geomecánicos",
    "🎲 Análisis UCS — Monte Carlo",
    "🪨 Granulometría",
])

# ╔══════════════════════════════════════════════╗
#  TAB 1 — DISEÑOS DE MEZCLA
# ╚══════════════════════════════════════════════╝
with tabs[0]:
    st.markdown("###  Diseños de Mezcla — Prueba Piloto")
    df        = st.session_state.designs_df.copy()
    resist_df = st.session_state.resist_df.copy()
    df["Total"]        = df["Cemento (kg/m³)"]+df["Agua (L/m³)"]+df["Desmonte (kg/m³)"]
    df["Cemento (%)"]  = (df["Cemento (kg/m³)"]/df["Total"]*100).round(2)
    df["Agua (%)"]     = (df["Agua (L/m³)"]/df["Total"]*100).round(2)
    df["Desmonte (%)"] = (df["Desmonte (kg/m³)"]/df["Total"]*100).round(2)
    df["A/C"]          = (df["Agua (L/m³)"]/df["Cemento (kg/m³)"]).round(3)

    st.markdown("#### Diseño de Mezcla (kg/m³ y L/m³)")
    st.markdown(render_table(df[["Diseño","Cemento (kg/m³)","Cemento (%)","Agua (L/m³)","Agua (%)",
                                  "Desmonte (kg/m³)","Desmonte (%)","Densidad (kg/m³)","A/C"]]),
                unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("#### Desarrollo de Resistencia a Compresión (MPa)")
    st.markdown(render_table(resist_df), unsafe_allow_html=True)
    st.session_state.designs_df = df

    st.markdown("### 🎯 Seleccionar Diseño")
    selected = st.selectbox("Seleccione un diseño:", df["Diseño"].tolist(), key="select_design")
    st.session_state.selected_design = selected
    st.info(f"📌 Diseño seleccionado: **{selected}**")

    st.markdown("###  Curva de Resistencia · Resistencia a 28 días")
    gc1, gc2 = st.columns(2)
    with gc1:
        fig_res = go.Figure()
        for d, color in zip(["D01","D02","D03","D04"], COLORS):
            fig_res.add_trace(go.Scatter(x=resist_df["Días"], y=resist_df[d],
                mode="lines+markers", name=d, opacity=1.0 if d==selected else 0.3,
                line=dict(color=color, width=4 if d==selected else 2)))
        fig_res.update_layout(**PLOTLY_LAYOUT, height=400,
            title=dict(text=f"Curva de Resistencia — {selected}", font=dict(size=15)),
            xaxis_title="Días", yaxis_title="MPa")
        st.plotly_chart(fig_res, use_container_width=True)
    with gc2:
        res_28 = resist_df[resist_df["Días"]==28].iloc[0]
        fig_28 = go.Figure()
        for d, color in zip(["D01","D02","D03","D04"], COLORS):
            fig_28.add_trace(go.Bar(x=[d], y=[res_28[d]], marker_color=color,
                opacity=1.0 if d==selected else 0.25,
                text=[f"{res_28[d]:.2f} MPa"], textposition="outside", textfont=dict(size=14)))
        fig_28.update_layout(**PLOTLY_LAYOUT, title=dict(text=f"Resistencia 28d — {selected}", font=dict(size=15)),
            height=400, showlegend=False, yaxis_title="MPa")
        st.plotly_chart(fig_28, use_container_width=True)

# ╔══════════════════════════════════════════════╗
#  TAB 2 — DISEÑO SELECCIONADO
# ╚══════════════════════════════════════════════╝
with tabs[1]:
    st.markdown("##  Diseño de Mezcla Seleccionado")
    if "selected_design" not in st.session_state:
        st.warning("Seleccione un diseño en la pestaña anterior."); st.stop()

    selected = st.session_state.selected_design
    st.markdown(f"### 🎯 DISEÑO: **{selected}**")
    df_all = st.session_state.designs_df.copy()
    if "A/C" not in df_all.columns:
        df_all["A/C"] = (df_all["Agua (L/m³)"]/df_all["Cemento (kg/m³)"]).round(3)
    df_resist = st.session_state.resist_df.copy()
    espec=st.session_state.espec; agre=st.session_state.agre
    row=df_all[df_all["Diseño"]==selected].iloc[0]
    cem=float(row["Cemento (kg/m³)"]); desm=float(row["Desmonte (kg/m³)"])
    rel_ac=float(row["A/C"]); resist_28=float(df_resist[df_resist["Días"]==28][selected].values[0])
    agua_ind=float(espec["Agua_industrial_Lm3"]); aire_at=int(espec["Aire_atrapado_pct"])
    hum=float(agre["Humedad_natural_pct"]); absorc=float(agre["Absorcion_pct"])
    peso_esp_cem=float(agre["Peso_especifico_cemento_kgm3"]); dens_agua=float(agre["Densidad_agua_kgm3"])
    agg_corr=desm/(1+hum/100); aporte=agg_corr*((hum-absorc)/100)
    dens_m=cem+agua_ind+aporte+agg_corr

    st.markdown("###  Especificaciones y Características del Agregado")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Especificaciones del Diseño**")
        st.markdown(render_table(pd.DataFrame({
            "Parámetro":["F'c (kg/cm²)","F'cr (kg/cm²)","Slump (pulg)","Agua ind. (L/m³)","Aire (%)","A/C"],
            "Valor":[f"{espec['Fc_kgcm2']:.0f}",f"{espec['Fcr_kgcm2']:.0f}",f"{espec['Slump_pulg']:.1f}",
                     f"{agua_ind:.1f}",f"{aire_at}",f"{rel_ac:.3f}"]
        })), unsafe_allow_html=True)
    with c2:
        st.markdown("**Características del Agregado**")
        st.markdown(render_table(pd.DataFrame({
            "Parámetro":["Tam. máx. (pulg)","Tam. nom. (pulg)","Mód. fineza",
                         "P.U. suelto (kg/m³)","P.U. compact. (kg/m³)","P.E. agr. (kg/m³)","Hum. nat. (%)","Absorción (%)"],
            "Valor":[f"{agre['Tamano_maximo_pulg']:.2f}",f"{agre['Tamano_nominal_pulg']:.2f}",f"{agre['Modulo_fineza']:.2f}",
                     f"{agre['Peso_unitario_suelto_kgm3']:.0f}",f"{agre['Peso_unitario_compactado_kgm3']:.0f}",
                     f"{agre['Peso_especifico_agregado_kgm3']:.0f}",f"{agre['Humedad_natural_pct']:.2f}",f"{agre['Absorcion_pct']:.2f}"]
        })), unsafe_allow_html=True)

    st.markdown("###  Adicionados y Cálculo de Mezcla")
    c3, c4 = st.columns(2)
    with c3:
        st.markdown("**Adicionados**")
        st.markdown(render_table(pd.DataFrame({
            "Propiedad":["P.E. cemento (kg/m³)","Densidad agua (kg/m³)"],
            "Valor":[f"{peso_esp_cem:.0f}",f"{dens_agua:.2f}"]
        })), unsafe_allow_html=True)
    with c4:
        st.markdown("**Cálculo del Diseño de Mezcla**")
        st.markdown(render_table(pd.DataFrame({
            "Concepto":["Cemento (kg/m³)","Agua ind. (L/m³)","Agg. corregido (kg/m³)",
                        "Aporte agua (L/m³)","Aire atrapado (%)","Densidad mezcla (kg/m³)"],
            "Valor":[f"{cem:.2f}",f"{agua_ind:.2f}",f"{agg_corr:.2f}",
                     f"{aporte:.2f}",f"{aire_at}%",f"{dens_m:.2f}"]
        })), unsafe_allow_html=True)

    st.markdown("### 🧮 Cálculo por Tanda")
    tanda = st.number_input("Tanda (m³ por ciclo)", value=10.0, step=0.5)
    st.markdown(render_table(pd.DataFrame({
        "Material":["Cemento","Agua (total)","Agregado"],
        "Por tanda":[f"{cem*tanda:.2f} kg",f"{(agua_ind+aporte)*tanda:.2f} L",f"{agg_corr*tanda:.2f} kg"]
    })), unsafe_allow_html=True)

    st.markdown("###  Gráficos del Diseño Seleccionado")
    cg1, cg2 = st.columns(2)
    with cg1:
        fig_pie = go.Figure(go.Pie(labels=["Cemento","Agua","Agregado"],
            values=[cem,agua_ind+aporte,agg_corr], hole=0.35, textfont=dict(size=14)))
        fig_pie.update_layout(**PLOTLY_LAYOUT, title=dict(text=f"Composición — {selected}", font=dict(size=15)))
        st.plotly_chart(fig_pie, use_container_width=True)
    with cg2:
        fig_bar = go.Figure(go.Bar(x=["Cemento","Agua total","Agregado"],
            y=[cem,agua_ind+aporte,agg_corr], marker_color=["#5fb840","#5898f8","#ffb830"],
            text=[f"{cem:.1f}",f"{agua_ind+aporte:.1f}",f"{agg_corr:.1f}"],
            textposition="outside", textfont=dict(size=14)))
        fig_bar.update_layout(**PLOTLY_LAYOUT,
            title=dict(text="Componentes por m³", font=dict(size=15)), yaxis_title="kg o L")
        st.plotly_chart(fig_bar, use_container_width=True)

    if st.button("💾 Guardar Diseño"):
        st.session_state.saved_mix = {"diseño":selected,"cemento":cem,"agua":agua_ind,
            "agregado":desm,"rel_ac":rel_ac,"densidad_mezcla":dens_m,"aire":aire_at,"resistencia_28d":resist_28}
        st.success("✔ Diseño guardado correctamente.")

# ╔══════════════════════════════════════════════╗
#  TAB 3 — INFRAESTRUCTURA
# ╚══════════════════════════════════════════════╝
with tabs[2]:
    st.markdown("##  Infraestructura · Cámara de Mezclado del CRF")
    if st.session_state.csv_activo:
        st.success("📂 Parámetros cargados desde el CSV maestro.")
    else:
        st.info("📋 Mostrando datos por defecto.")

    infra = st.session_state.infra
    largo=float(infra["Largo_camara_m"]); ancho=float(infra["Ancho_util_m"]); alto=float(infra["Alto_util_m"])
    esp_conc=float(infra["Espesor_concreto_m"]); esp_shot=float(infra["Espesor_shotcrete_m"])
    esp_rell=float(infra["Espesor_relleno_estructural_m"])
    alt_buz=float(infra["Altura_buzon_m"]); anc_ban=float(infra["Ancho_bandeja_m"])

    pendiente = st.selectbox("Pendiente del piso (%)", [-1, -2])
    st.session_state.pendiente = pendiente

    st.markdown("###  Equipos y Personal")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Equipos Principales**")
        st.markdown(render_table(pd.DataFrame({
            "Equipo":["Bomba estacionaria","Chancadora secundaria","Faja transportadora","Tolva de regulación"],
            "Capacidad":[f"{infra['Bomba_estacionaria_m3h']:.0f} m³/h",f"{infra['Chancadora_secundaria_th']:.0f} t/h",
                         f"{infra['Faja_transportadora_th']:.0f} t/h",f"{infra['Tolva_regulacion_m3']:.0f} m³"]
        })), unsafe_allow_html=True)
    with c2:
        st.markdown("**Personal Directo Requerido**")
        st.markdown(render_table(pd.DataFrame({
            "Cargo":["Operador","Mecánico","Supervisor"],
            "Cantidad":[int(infra["Personal_operadores"]),int(infra["Personal_mecanicos"]),int(infra["Personal_supervisores"])]
        })), unsafe_allow_html=True)

    piso_drop=largo*(abs(pendiente)/100)
    vol_util=largo*ancho*alto; vol_conc=largo*ancho*esp_conc
    vol_exc=(ancho+2*esp_conc)*(alto+2*esp_conc)*largo

    st.markdown("###  Geometría y Volúmenes")
    c3, c4 = st.columns(2)
    with c3:
        st.markdown("**Geometría de la Cámara**")
        st.markdown(render_table(pd.DataFrame({
            "Parámetro":["Longitud (m)","Ancho útil (m)","Altura útil (m)","Pendiente piso (%)",
                         "Esp. concreto (m)","Esp. shotcrete (m)","Esp. relleno (m)","Altura buzón (m)","Ancho bandeja (m)"],
            "Valor":[largo,ancho,alto,pendiente,esp_conc,esp_shot,esp_rell,alt_buz,anc_ban]
        })), unsafe_allow_html=True)
    with c4:
        st.markdown("**Cálculos de Volúmenes**")
        st.markdown(render_table(pd.DataFrame({
            "Concepto":["Volumen útil (m³)","Caída del piso (m)","Vol. concreto (m³)","Vol. excavado (m³)"],
            "Valor":[f"{vol_util:.2f}",f"{piso_drop:.2f}",f"{vol_conc:.2f}",f"{vol_exc:.2f}"]
        })), unsafe_allow_html=True)

    st.markdown("##  Perfil Técnico — Cámara de Mezclado CRF")
    fig_tec = draw_mining_profile(largo, ancho, alto, esp_conc, esp_shot, esp_rell, alt_buz, anc_ban, pendiente)
    st.pyplot(fig_tec, use_container_width=True)
    plt.close(fig_tec)

# ╔══════════════════════════════════════════════╗
#  TAB 4 — CÁLCULOS GEOMECÁNICOS  (sin opción PDF)
# ╚══════════════════════════════════════════════╝
with tabs[3]:
    st.markdown("##  Cálculos Geomecánicos — Método Mitchell + Dujisin")
    if st.session_state.csv_activo:
        st.success("📂 Parámetros cargados desde el CSV maestro.")
    else:
        st.info("📋 Mostrando datos por defecto.")

    geo=st.session_state.geo
    H=float(geo["Altura_caseron_H_m"]); B=float(geo["Ancho_caseron_B_m"]); L=float(geo["Longitud_caseron_L_m"])
    dens_rell=float(geo["Densidad_CRF_kgm3"]); cohes=float(geo["Cohesion_CRF_kPa"])
    phi=float(geo["Angulo_friccion_phi_deg"]); dens_roca=float(geo["Densidad_roca_kgm3"])
    fs=float(geo["Factor_seguridad"]); poisson=float(geo["Coeficiente_Poisson"])
    gamma_rell=dens_rell*9.81/1e6; gamma_roca=dens_roca*9.81/1e6
    K=(1-math.sin(math.radians(phi)))/(1+math.sin(math.radians(phi)))
    sigma_v=gamma_rell*H; sigma_h=K*sigma_v
    sigma_techo=sigma_v/fs; sigma_lat=sigma_h/fs; sigma_tot=sigma_techo+sigma_lat

    st.markdown("###  Parámetros de Entrada")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Geometría y Materiales**")
        st.markdown(render_table(pd.DataFrame({
            "Parámetro":["Altura caserón H (m)","Ancho caserón B (m)","Longitud caserón L (m)",
                         "Densidad CRF (kg/m³)","Cohesión CRF (kPa)"],
            "Valor":[f"{H:.1f}",f"{B:.1f}",f"{L:.1f}",f"{dens_rell:.0f}",f"{cohes:.0f}"]
        })), unsafe_allow_html=True)
    with c2:
        st.markdown("**Propiedades Geomecánicas**")
        st.markdown(render_table(pd.DataFrame({
            "Parámetro":["Ángulo fricción φ (°)","Densidad roca (kg/m³)","Factor de Seguridad","Coef. Poisson ν"],
            "Valor":[f"{phi:.1f}",f"{dens_roca:.0f}",f"{fs}",f"{poisson:.2f}"]
        })), unsafe_allow_html=True)

    st.markdown("###  Conversión y Mitchell (1970)")
    c3, c4 = st.columns(2)
    with c3:
        st.markdown("**Pesos Unitarios**")
        st.markdown(render_table(pd.DataFrame({
            "Material":["Relleno CRF","Roca encajonante"],
            "Densidad (kg/m³)":[dens_rell,dens_roca],
            "γ (MN/m³)":[f"{gamma_rell:.6f}",f"{gamma_roca:.6f}"]
        })), unsafe_allow_html=True)
    with c4:
        st.markdown("**Cálculos Mitchell (1970)**")
        st.markdown(render_table(pd.DataFrame({
            "Concepto":["Sobrecarga vertical σv (MN/m²)","Coef. activo K","Presión lateral σh (MN/m²)"],
            "Valor":[f"{sigma_v:.4f}",f"{K:.4f}",f"{sigma_h:.4f}"]
        })), unsafe_allow_html=True)

    st.markdown("###  Dujisin (1992) y Resultado")
    c5, c6 = st.columns(2)
    with c5:
        st.markdown("**Presiones Reducidas**")
        st.markdown(render_table(pd.DataFrame({
            "Concepto":["σ techo reducida (MN/m²)","σ lateral reducida (MN/m²)","Factor de Seguridad"],
            "Valor":[f"{sigma_techo:.4f}",f"{sigma_lat:.4f}",f"{fs}"]
        })), unsafe_allow_html=True)
    with c6:
        st.markdown("**Presión Total Actuante**")
        st.markdown(render_table(pd.DataFrame({
            "Concepto":["σ total sobre caserón (MN/m²)"],
            "Valor":[f"{sigma_tot:.4f}"]
        })), unsafe_allow_html=True)

    st.info(f"📌 **Presión total actuante sobre el caserón:** {sigma_tot:.4f} MN/m²")

    st.markdown("###  Gráficos Geomecánicos")
    cg1, cg2 = st.columns(2)
    z = np.linspace(0, H, 50)
    with cg1:
        fig1 = go.Figure()
        fig1.add_trace(go.Scatter(x=K*gamma_rell*z, y=z, name="Mitchell", mode="lines", line=dict(color="#ffb830",width=3)))
        fig1.add_trace(go.Scatter(x=K*gamma_rell*z/fs, y=z, name="Dujisin", mode="lines", line=dict(color="#5fb840",width=3,dash="dash")))
        fig1.update_layout(**PLOTLY_LAYOUT, title=dict(text="Presión Lateral (MN/m²)", font=dict(size=15)),
            xaxis_title="Presión", yaxis_title="Altura (m)", height=400)
        st.plotly_chart(fig1, use_container_width=True)
    with cg2:
        fig2 = go.Figure(go.Bar(x=["σ techo reducida"], y=[sigma_techo],
            text=[f"{sigma_techo:.4f} MN/m²"], textposition="outside",
            textfont=dict(size=14), marker_color="#5898f8"))
        fig2.update_layout(**PLOTLY_LAYOUT, title=dict(text="Soporte Requerido — Techo", font=dict(size=15)),
            height=400, showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)

# ╔══════════════════════════════════════════════╗
#  TAB 5 — ANÁLISIS UCS MONTE CARLO  (posición 5, antes de granulometría)
# ╚══════════════════════════════════════════════╝
with tabs[4]:
    st.markdown("## 🎲 Análisis de UCS — Métodos Determinísticos + Simulación Monte Carlo")
    st.markdown("""
    <div style="background:#2a1808;border-left:4px solid #ffb830;padding:12px 18px;border-radius:4px;margin-bottom:14px;">
    <b style="color:#ffb830;font-size:15px;">Métodos implementados:</b>
    <span style="color:#f0d8b8;font-size:15px;"> Yu (2006) · Potvin (2005) · Mitchell (1982) · Obert y Duvall (1967) — 
    Análisis estocástico con distribuciones truncadas normales</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("### ⚙️ Parámetros de Diseño para UCS")
    p = {}
    row1 = st.columns(4)
    row2 = st.columns(4)
    all_cols = row1 + row2
    for i, (k, v) in enumerate(PARAM_INFO.items()):
        p[k] = all_cols[i].number_input(
            f"{v['label']} [{v['unit']}]",
            min_value=float(v["min"]), max_value=float(v["max"]),
            value=float(st.session_state.mc_params.get(k, v["default"])),
            step=float(v["step"]), help=v["hint"], key=f"mc_{k}")
    st.session_state.mc_params = p

    col_n, col_btn = st.columns([2, 1])
    with col_n:
        n_mc = st.select_slider("Simulaciones Monte Carlo",
            options=[1000, 5000, 10000, 25000, 50000], value=10000)
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True)
        run_btn = st.button("▶  Calcular UCS + Monte Carlo", type="primary", use_container_width=True)

    if run_btn:
        with st.spinner("Ejecutando simulación Monte Carlo..."):
            det = {
                "UCS_Yu":       calc_ucs_yu(p["gamma"], p["H"]),
                "UCS_Potvin":   calc_ucs_potvin(p["gamma"], p["B"], p["H"], p["D"]),
                "UCS_Mitchell": calc_ucs_mitchell(p["gamma"], p["B"], p["H"], p["D"], p["phi"]),
                "UCS_Techo":    calc_ucs_roof(p["gamma"], p["L"], p["t"], p["FS"]),
            }
            inputs, outputs, keys = monte_carlo(p, n_mc)
            mc_stats_list = [mc_stats(outputs[:, i]) for i in range(4)]
            _, corr = fig_mc_tornado(inputs, outputs, keys)
            st.session_state.mc_results = dict(
                det=det, inputs=inputs, outputs=outputs,
                keys=keys, mc_stats=mc_stats_list, corr=corr, p=p
            )
            st.success(f"✔ Simulación completada — {n_mc:,} iteraciones.")

    if st.session_state.mc_results:
        r = st.session_state.mc_results
        det, inputs, outputs, keys, mc_stats_list, corr = (
            r["det"], r["inputs"], r["outputs"], r["keys"], r["mc_stats"], r["corr"])

        st.markdown("### 📌 Resultados Determinísticos")
        cols = st.columns(4)
        for col, (name, val) in zip(cols, det.items()):
            label, color = interpret_ucs(val)
            col.metric(label=name.replace("UCS_",""), value=f"{val:.4f} MPa", delta=label)

        st.divider()

        mct1, mct2, mct3, mct4, mct5 = st.tabs(
            ["📊 Histogramas","📈 CDF","🌪 Tornado","✦ Dispersión","📦 Box Plot"])

        with mct1:
            st.markdown("**Distribución de frecuencias Monte Carlo** — densidad de probabilidad por método")
            fig = fig_mc_histograms(outputs)
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

        with mct2:
            st.markdown("**Función de distribución acumulada** — probabilidad de excedencia")
            fig = fig_mc_cdf(outputs)
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

        with mct3:
            st.markdown("**Sensibilidad de parámetros** respecto a UCS Yu — Correlación de Pearson")
            fig_t, _ = fig_mc_tornado(inputs, outputs, keys)
            st.pyplot(fig_t, use_container_width=True)
            plt.close(fig_t)

        with mct4:
            st.markdown("**Diagramas de dispersión** — cada variable de entrada vs UCS Yu")
            fig = fig_mc_scatter(inputs, outputs, keys)
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

        with mct5:
            st.markdown("**Box-plot comparativo** de las cuatro distribuciones Monte Carlo")
            fig = fig_mc_box(outputs)
            st.pyplot(fig, use_container_width=True)
            plt.close(fig)

        st.divider()

        st.markdown("### 📋 Estadísticas Monte Carlo por Método")
        rows = []
        for i, name in enumerate(METHOD_NAMES):
            s = mc_stats_list[i]
            rows.append({"Método":name, "Media":f"{s['mean']:.4f}","Std":f"{s['std']:.4f}",
                         "P5":f"{s['p5']:.4f}","P50":f"{s['p50']:.4f}","P90":f"{s['p90']:.4f}","P95":f"{s['p95']:.4f}"})
        st.markdown(render_table(pd.DataFrame(rows)), unsafe_allow_html=True)

        st.divider()

        st.markdown("### 📥 Exportar Resultados Monte Carlo")
        excel_bytes = export_mc_excel(p, det, mc_stats_list, inputs, outputs, corr, keys)
        st.download_button(
            label="⬇  Descargar Excel completo (UCS + Monte Carlo)",
            data=excel_bytes,
            file_name=f"UCS_MonteCarlo_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("Incluye: Datos de entrada · Interpretación · Estadísticas MC · Sensibilidad · Muestra MC (1,000 filas)")
    else:
        st.markdown("""
        <div style='text-align:center; padding:60px 20px; color:#c9a87c;'>
          <div style='font-size:3.5rem;'>⛏️</div>
          <h3 style='color:#ffb830;'>Configure los parámetros y presione <em>Calcular</em></h3>
          <p style='max-width:480px; margin:0 auto; line-height:1.6; color:#f0d8b8;'>
            Ingrese los valores geotécnicos y ejecute el análisis determinístico + simulación Monte Carlo
            para obtener distribuciones de UCS, tornado de sensibilidad y estadísticas de percentiles.
          </p>
        </div>
        """, unsafe_allow_html=True)

# ╔══════════════════════════════════════════════╗
#  TAB 6 — GRANULOMETRÍA (última pestaña — incluye botón PDF)
# ╚══════════════════════════════════════════════╝
with tabs[5]:
    st.markdown("## 🪨 Análisis Granulométrico — ASTM D6913 / D2487 / D1140")
    st.markdown("""
    <div style="background:#2a1808;border-left:4px solid #ffb830;padding:12px 18px;border-radius:4px;margin-bottom:14px;">
    <b style="color:#ffb830;font-size:15px;">Normas Aplicadas:</b>
    <span style="color:#f0d8b8;font-size:15px;"> ASTM D6913-17 (Análisis Granulométrico por Tamices) · ASTM D2487-17 (USCS Clasificación) · ASTM D1140-00 (Finos pasante N°200)</span>
    </div>
    """, unsafe_allow_html=True)

    agre = st.session_state.agre
    tam_max_pulg = agre["Tamano_maximo_pulg"]
    tam_nom_pulg = agre["Tamano_nominal_pulg"]
    mf_agre = agre["Modulo_fineza"]

    st.markdown("### 📥 Ingreso de Datos Granulométricos")
    st.info(f"Datos del agregado cargados: Tam. máx = **{tam_max_pulg}\"** | Tam. nom. = **{tam_nom_pulg}\"** | Mód. fineza = **{mf_agre}**")

    col_opt1, col_opt2 = st.columns([2, 1])
    with col_opt1:
        input_mode = st.radio("Modo de ingreso de datos granulométricos:",
            ["Automático (basado en parámetros del agregado cargado)",
             "Manual (ingresar % retenido por tamiz)"],
            horizontal=True)
    with col_opt2:
        masa_total = st.number_input("Masa total de muestra (g)", value=2000.0, step=100.0, min_value=50.0,
                                      help="ASTM D6913 — mínimo según tamaño máximo de partícula")

    if input_mode.startswith("Automático") and st.session_state.gran_retained is None:
        retained_input = default_gradacion_CRF(agre)
        st.info("⚙️ Curva generada automáticamente desde parámetros del agregado. Puede ajustarla cambiando al modo Manual.")
    elif st.session_state.gran_retained is not None:
        retained_input = st.session_state.gran_retained
    else:
        retained_input = default_gradacion_CRF(agre)

    if input_mode.startswith("Manual"):
        st.markdown("#### Ingrese % Retenido Parcial por Tamiz (ASTM D6913 — Tabla 1)")
        st.caption("Ingrese el porcentaje de masa retenida en cada tamiz. La suma debe ser ≤ 100%.")
        col_grid = st.columns(4)
        manual_vals = []
        for idx, (sname, smm) in enumerate(zip(SIEVES_NAME, SIEVES_MM)):
            default_v = retained_input[idx] if idx < len(retained_input) else 0.0
            col = col_grid[idx % 4]
            v = col.number_input(f"{sname} ({smm}mm)", value=float(default_v),
                                  min_value=0.0, max_value=100.0, step=0.1,
                                  key=f"gran_{idx}")
            manual_vals.append(v)
        retained_input = manual_vals
        st.session_state.gran_retained = retained_input

    gd = calc_granulometria(retained_input)
    st.session_state.gran_calc = gd

    st.markdown("### 📊 Resultados del Análisis — ASTM D6913 / D2487")
    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Grava ≥ 4.75mm", f"{gd['grava_pct']:.1f}%")
    m2.metric("Arena 0.075–4.75mm", f"{gd['arena_pct']:.1f}%")
    m3.metric("Finos < 0.075mm", f"{gd['finos_pct']:.1f}%")
    m4.metric("Mód. Fineza (MF)", f"{gd['modulo_fineza']:.2f}")
    m5.metric("Clasificación USCS", gd['clasificacion'].split(" ")[0])

    st.markdown("---")

    st.markdown("### 📋 Tabla Granulométrica Detallada (ASTM D6913)")
    masa_retenida = [p/100.0 * masa_total for p in gd["retained_pct"]]
    masa_ret_acum = np.cumsum(masa_retenida)
    df_gran_full = pd.DataFrame({
        "Tamiz": gd["sieves_name"],
        "Abertura (mm)": [f"{mm:.3f}" for mm in gd["sieves_mm"]],
        "Masa Ret. (g)": [f"{m:.1f}" for m in masa_retenida],
        "Masa Ret. Acum. (g)": [f"{m:.1f}" for m in masa_ret_acum],
        "% Ret. Parcial": [f"{v:.2f}" for v in gd["retained_pct"]],
        "% Ret. Acumulado": [f"{v:.2f}" for v in gd["cum_ret"]],
        "% Que Pasa": [f"{v:.2f}" for v in gd["pct_pasa"]],
    })
    st.markdown(render_table(df_gran_full), unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 🔬 Parámetros Granulométricos — ASTM D2487")
    pc1, pc2 = st.columns(2)
    with pc1:
        d10_s = f"{gd['D10']:.4f} mm" if gd['D10'] else "N/A (extrapolar)"
        d30_s = f"{gd['D30']:.4f} mm" if gd['D30'] else "N/A"
        d60_s = f"{gd['D60']:.4f} mm" if gd['D60'] else "N/A (extrapolar)"
        cu_s  = f"{gd['Cu']:.2f}" if gd['Cu'] else "N/A"
        cc_s  = f"{gd['Cc']:.3f}" if gd['Cc'] else "N/A"
        st.markdown(render_table(pd.DataFrame({
            "Parámetro":["D₁₀ (diámetro efectivo)","D₃₀","D₆₀ (diámetro controlante)",
                         "Cu = D₆₀/D₁₀ (Coef. Uniformidad)","Cc = D₃₀²/(D₁₀·D₆₀) (Coef. Curvatura)"],
            "Valor": [d10_s, d30_s, d60_s, cu_s, cc_s],
            "Criterio ASTM D2487": [
                "Valor de referencia",
                "Valor de referencia",
                "Valor de referencia",
                "GW: Cu≥4 | SW: Cu≥6",
                "Bien grad.: 1≤Cc≤3",
            ]
        })), unsafe_allow_html=True)
    with pc2:
        if gd['Cu'] and gd['Cc']:
            gw_ok = "✅ Cumple" if gd['Cu'] >= 4 else "❌ No cumple"
            sw_ok = "✅ Cumple" if gd['Cu'] >= 6 else "❌ No cumple"
            cc_ok = "✅ Cumple (1≤Cc≤3)" if (gd['Cc'] and 1 <= gd['Cc'] <= 3) else "❌ No cumple"
        else:
            gw_ok = sw_ok = cc_ok = "N/A"
        st.markdown(render_table(pd.DataFrame({
            "Verificación USCS":["Clasificación final","Grava %","Arena %","Finos %",
                                 "Cu≥4 (GW)","Cu≥6 (SW)","1≤Cc≤3"],
            "Resultado":[gd['clasificacion'], f"{gd['grava_pct']:.1f}%",
                         f"{gd['arena_pct']:.1f}%", f"{gd['finos_pct']:.1f}%",
                         gw_ok, sw_ok, cc_ok]
        })), unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📈 Curva Granulométrica — ASTM D6913")
    fig_gran = go.Figure()
    fig_gran.add_trace(go.Scatter(
        x=gd["sieves_mm"], y=gd["pct_pasa"],
        mode="lines+markers", name="Curva Granulométrica CRF",
        line=dict(color="#ffb830", width=3),
        marker=dict(size=8, symbol="circle", color="#ffb830",
                    line=dict(color="#fff", width=1)),
        hovertemplate="<b>%{text}</b><br>Abertura: %{x:.3f} mm<br>% Pasa: %{y:.1f}%<extra></extra>",
        text=gd["sieves_name"]
    ))
    for d_val, d_label, col_v in [
        (gd["D10"], "D₁₀", "#5fb840"),
        (gd["D30"], "D₃₀", "#5898f8"),
        (gd["D60"], "D₆₀", "#f85149")
    ]:
        if d_val:
            fig_gran.add_vline(x=d_val, line_dash="dash", line_color=col_v, line_width=1.5, opacity=0.8)
            fig_gran.add_annotation(x=math.log10(d_val), y=8, text=f"<b>{d_label}</b><br>{d_val:.3f}mm",
                                     xref="x", yref="y", showarrow=False,
                                     font=dict(color=col_v, size=10),
                                     bgcolor="#1a1008", bordercolor=col_v, borderwidth=1)
    fig_gran.add_vline(x=0.075, line_dash="dot", line_color="#8b7355", line_width=1, opacity=0.6)
    fig_gran.add_vline(x=4.75,  line_dash="dot", line_color="#8b7355", line_width=1, opacity=0.6)
    fig_gran.add_annotation(x=math.log10(0.035), y=90, text="Finos", xref="x", yref="y",
                             showarrow=False, font=dict(color="#c9a87c", size=11))
    fig_gran.add_annotation(x=math.log10(0.5), y=90, text="Arena", xref="x", yref="y",
                             showarrow=False, font=dict(color="#c9a87c", size=11))
    fig_gran.add_annotation(x=math.log10(15), y=90, text="Grava", xref="x", yref="y",
                             showarrow=False, font=dict(color="#c9a87c", size=11))

    fig_gran.update_layout(
        **PLOTLY_LAYOUT,
        height=500,
        title=dict(text="Curva Granulométrica de Distribución de Tamaño de Partícula (ASTM D6913)", font=dict(size=14))
    )
    st.plotly_chart(fig_gran, use_container_width=True)

    st.markdown("### 📊 Composición por Fracción y % Retenido Parcial")
    gc1, gc2 = st.columns(2)
    with gc1:
        fig_frac = go.Figure(go.Pie(
            labels=["Grava (≥4.75mm)", "Arena (0.075–4.75mm)", f"Finos (<0.075mm)"],
            values=[gd["grava_pct"], gd["arena_pct"], gd["finos_pct"]],
            hole=0.4,
            marker=dict(colors=["#ffb830", "#5fb840", "#5898f8"]),
            textfont=dict(size=13)
        ))
        fig_frac.update_layout(**PLOTLY_LAYOUT, height=380,
            title=dict(text="Distribución por Fracción (ASTM D2487)", font=dict(size=13)))
        st.plotly_chart(fig_frac, use_container_width=True)
    with gc2:
        fig_ret = go.Figure(go.Bar(
            x=gd["sieves_name"], y=gd["retained_pct"],
            marker_color=["#ffb830" if mm >= 4.75 else "#5fb840" if mm >= 0.075 else "#5898f8"
                          for mm in gd["sieves_mm"]],
            text=[f"{v:.1f}%" for v in gd["retained_pct"]],
            textposition="outside", textfont=dict(size=10)
        ))
        fig_ret.update_layout(**PLOTLY_LAYOUT, height=380,
            title=dict(text="% Retenido Parcial por Tamiz", font=dict(size=13))
            )
        st.plotly_chart(fig_ret, use_container_width=True)

    st.markdown("---")
    st.markdown("### 🔍 Verificación Finos — ASTM D1140")
    finos_d1140 = gd["pct_pasa"][-1]
    col_d1, col_d2, col_d3 = st.columns(3)
    col_d1.metric("% Pasante N°200 (por lavado)", f"{finos_d1140:.2f}%", help="ASTM D1140 — lavado en húmedo")
    masa_finos = finos_d1140/100 * masa_total
    col_d2.metric("Masa de finos (g)", f"{masa_finos:.1f} g")
    clasificacion_finos = "Fino" if finos_d1140 >= 50 else ("Limo/Arcilla" if finos_d1140 >= 12 else "Limpio" if finos_d1140 < 5 else "Con finos")
    col_d3.metric("Clasificación preliminar", clasificacion_finos)
    st.markdown(f"""
    <div style="background:#2a1808;padding:12px 18px;border-radius:4px;border-left:3px solid #5fb840;">
    <b style="color:#ffb830;font-size:15px;">Interpretación ASTM D1140 / D2487:</b><br>
    <span style="color:#f0d8b8;font-size:15px;">
    — Finos pasante N°200: <b>{finos_d1140:.2f}%</b> · Masa lavada: <b>{masa_finos:.1f} g</b><br>
    — {'⚠️ Suelo de grano fino — se requieren ensayos de Límites de Atterberg (D4318) para clasificación USCS completa.' 
       if finos_d1140 >= 50 else 
       '✅ Fracción gruesa predominante — clasificación USCS basada en granulometría (Cu y Cc).'}<br>
    — Clasificación USCS estimada: <b>{gd["clasificacion"]}</b>
    </span></div>""", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### ⚖️ Verificación vs Parámetros del Agregado Declarados")
    mf_calculado = gd["modulo_fineza"]
    mf_diff = abs(mf_calculado - mf_agre)
    diff_pct = mf_diff / mf_agre * 100 if mf_agre > 0 else 0
    col_v1, col_v2, col_v3 = st.columns(3)
    col_v1.metric("MF Declarado", f"{mf_agre:.2f}")
    col_v2.metric("MF Calculado (ASTM)", f"{mf_calculado:.2f}", delta=f"{mf_calculado-mf_agre:+.2f}")
    col_v3.metric("Diferencia", f"{diff_pct:.1f}%",
                  delta="✅ Consistente" if diff_pct < 5 else "⚠️ Revisar",
                  delta_color="normal" if diff_pct < 5 else "inverse")

    # ──────────────────────────────────────────
    #  REPORTE PDF — ahora en la última pestaña
    # ──────────────────────────────────────────
    st.markdown("---")
    st.markdown("""
    <div style="background:#2a1808;border:2px solid #ffb830;border-radius:8px;padding:16px 20px;margin-top:10px;">
    <h3 style="color:#ffb830;margin-top:0;">📄 Exportar Reporte Completo PDF</h3>
    <p style="color:#f0d8b8;font-size:15px;margin-bottom:0;">
    El reporte incluye todos los módulos: Diseños de Mezcla · Diseño Seleccionado · Infraestructura · 
    Cálculos Geomecánicos · Granulometría · Análisis UCS Monte Carlo.
    </p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

    if "selected_design" not in st.session_state:
        st.warning("⚠️ Seleccione un diseño en la pestaña 'Diseños de Mezcla' antes de generar el reporte.")
    else:
        if st.button("📄 Generar Reporte PDF Completo", type="primary", use_container_width=True):
            with st.spinner("Generando reporte PDF completo (CRF + Granulometría + Monte Carlo)..."):
                try:
                    gran_data = None
                    if st.session_state.gran_retained:
                        gran_data = calc_granulometria(st.session_state.gran_retained)
                    elif st.session_state.get("gran_calc"):
                        gran_data = st.session_state.gran_calc

                    pdf_bytes = build_pdf_report({
                        "selected":   st.session_state.selected_design,
                        "designs_df": st.session_state.designs_df,
                        "resist_df":  st.session_state.resist_df,
                        "espec":      st.session_state.espec,
                        "agre":       st.session_state.agre,
                        "infra":      st.session_state.infra,
                        "geo":        st.session_state.geo,
                        "pendiente":  st.session_state.get("pendiente", -1),
                        "gran_data":  gran_data,
                        "mc_results": st.session_state.mc_results,
                    })
                    hoy_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")
                    st.download_button(
                        label="⬇️ Descargar Reporte PDF Completo",
                        data=pdf_bytes,
                        file_name=f"Reporte_CRF_Completo_{st.session_state.selected_design}_{hoy_str}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                    )
                    st.success("✔ Reporte completo generado — incluye Granulometría y Monte Carlo.")
                except Exception as e:
                    st.error(f"❌ Error al generar el reporte: {e}")
                    st.exception(e)

# ─────────────────────────────────────────────
#  FOOTER
# ─────────────────────────────────────────────
hoy = datetime.datetime.now().strftime("%d/%m/%Y")
st.markdown(f"""
<style>
.footer-box {{
    margin-top: 40px; padding: 20px; width: 100%; text-align: center;
    background: rgba(255,184,48,0.08); border-top: 2px solid #ffb830;
    border-radius: 6px; font-size: 15px; color: #f5e6d0;
}}
.footer-title {{ font-family: 'Lato', sans-serif; font-size: 17px; color: #ffb830; font-weight: 900; }}
.footer-sub   {{ color: #c9a87c; font-size: 14px; }}
</style>
<div class="footer-box">
    <div class="footer-title">Cemented Rock Fill · Sistema CRF Integrado</div>
    <div class="footer-sub">CRF · Granulometría ASTM D6913/D2487/D1140 · Análisis UCS Monte Carlo · Sub Level Stoping</div>
    <br>
    <div class="footer-sub">Versión 2.0 · Actualizado el {hoy}</div>
    <div class="footer-sub">Desarrollado por: <b>Bradoc Chambilla</b></div>
</div>
""", unsafe_allow_html=True)
